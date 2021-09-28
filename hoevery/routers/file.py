from fastapi import FastAPI, File, UploadFile, Form
from config import settings
from fastapi import APIRouter, Request, Response, Depends
import shutil
from typing import List
from config.db import *
from fastapi.responses import FileResponse
from starlette.responses import StreamingResponse

from routers.import_excel import *

#import pandas as pd
import openpyxl
#import xlrd
#import xlsxwriter
import pyexcel as p
from io import BytesIO

router = APIRouter()

def rqp2dict(request):
    result = {}
    for d in request.query_params._list:
        if result.get(d[0]):
            if type(result.get(d[0])) is list:
                result[d[0]] = result[d[0]] + [d[1]]
            else:
                result[d[0]] = [result[d[0]], d[1]]
        else:
            result[d[0]] = d[1]
    return result


@router.post("/upload-file")
async def create_upload_file(username:str = "username_file", uploaded_file: UploadFile = File(...)):
    file_location = f"files/upload/{username}_{uploaded_file.filename}"
    with open(file_location, "wb+") as file_object:
        file_object.write(uploaded_file.file.read())
    return dict(ret=0, msg="Complete.", data={"info": f"file '{uploaded_file.filename}' saved at '{file_location}'"})


@router.post("/upload-multi-file")
async def upload_file(username:str = "username_file", files: List[UploadFile] = File(...)):
    list = []
    for img in files:
        file_location = f"files/upload/{username}_{img.filename}"
        list.append(f"{username}_{img.filename}")
        with open(file_location, "wb") as buffer:
            shutil.copyfileobj(img.file, buffer)
    return dict(ret=0, msg="Complete.", data={"info": f"files '{list}' saved at 'files/upload/'"})


@router.get("/download-file/{path}")
async def download_file(path: str):
    return FileResponse(f"files/upload/{path}", media_type='application/octet-stream',filename=f"{path}")


# @router.get("/export-excel")
# async def export_excel():
#     return "ok"


#@router.post("/import-excel")
async def import_excel(file: UploadFile = File(...)):
    # upload file xlrd before read file.
    file_location = f"files/upload/{file.filename}"
    with open(file_location, "wb+") as file_object:
        file_object.write(file.file.read())

    # To open the workbook, workbook object is created
    wb_obj = openpyxl.load_workbook(file_location)

    # Get workbook active sheet object from the active attribute
    sheet_obj = wb_obj.active

    # Note: The first row or column integer is 1, not 0.
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    if (max_col == 2):    
        result = import_company(max_col, max_row, sheet_obj)
    if (max_col == 3):    
        result = import_department(max_col, max_row, sheet_obj)
    if (max_col == 6):    
        result = import_employee(max_col, max_row, sheet_obj)
    
    return result

#@router.get("/export-excel-company")
async def export_excel(request: Request):
    file_location = f'files/upload/report_company.xls'
    with SessionContext() as se:
        query = db.company()
        params = rqp2dict(request)
        for key, value in params.items():
            try:
                query = query.filter(
                    getattr(db.company, key) == value)
            except Exception as e:
                print(key, "not found")
        
        total = query.count()
        records_db = query.all()
        records = []
        for item in records_db:
            records.append({"id": item.id, "name": item.name, "description": item.description})
        p.save_as(records=records, dest_file_name=file_location)  
    return FileResponse(file_location, media_type='application/octet-stream',filename='report_company.xls')
    # return dict(ret=0, msg="Complete", data1=records_db)


#@router.get("/export-excel-department")
async def export_excel(request: Request):
    file_location = f'files/upload/report_department.xls'
    with SessionContext() as se:
        query = db.department()
        params = rqp2dict(request)
        for key, value in params.items():
            try:
                query = query.filter(
                    getattr(db.department, key) == value)
            except Exception as e:
                print(key, "not found")
        
        total = query.count()
        records_db = query.all()
        records = []
        for item in records_db:
            records.append({"id": item.id, "corp_id": item.corp_id, "name": item.name, "description": item.description})
        p.save_as(records=records, dest_file_name=file_location) 
    return FileResponse(file_location, media_type='application/octet-stream',filename='report_department.xls') 

#@router.get("/export-excel-employee")
async def export_excel(request: Request):
    file_location = f'files/upload/report_employee.xls'
    with SessionContext() as se:
        query = db.employee()
        params = rqp2dict(request)
        for key, value in params.items():
            try:
                query = query.filter(
                    getattr(db.employee, key) == value)
            except Exception as e:
                print(key, "not found")
        
        total = query.count()
        records_db = query.all()
        records = []
        for item in records_db:
            records.append({"id": item.id, "corp_id": item.corp_id, "dept_id": item.dept_id, 
                            "firstname": item.firstname, "lastname": item.lastname, "emp_code": item.emp_code,
                            "birthdate": item.birthdate, "create_at": item.create_at, "create_by": item.create_by,
                            "update_at": item.update_at, "update_by": item.update_by})
        p.save_as(records=records, dest_file_name=file_location) 
    return FileResponse(file_location, media_type='application/octet-stream',filename='report_employee.xls')

# @router.post("impot-excel-pyexcel")
async def import_excel_pyexcel(file: UploadFile = File(...)):
    # upload file xlrd before read file.
    file_location = f"files/upload/{file.filename}"
    with open(file_location, "wb+") as file_object:
        file_object.write(file.file.read())

    _dict = {}
    success = 0
    with SessionContext() as se: # insert into db 
        newcompany = db.company()
        for row in p.get_array(file_name=file_location, start_row=1):
            _dict['company'] = row[0]
            _dict['name'] = row[1]
            _dict['description'] = row[2]
            try:
                newcompany = se.query(db.company).filter(db.company.name == row[1]).first()
                if newcompany:
                    error = row[1] + " already exits"
                    return dict(ret=-1, msg="Failed", total="total", success=success, error=dict(row=success+1, detail=error))
            except Exception as e:
                return dict(ret=-1, msg="Not match")

            newcompany = db.company()
            for key in _dict:
                print(newcompany, key)
                if hasattr(newcompany, key):
                    setattr(newcompany, key, _dict.get(key))

            se.add(newcompany)
            se.commit()
            se.refresh(newcompany)
            success += 1

    return dict(ret=0, msg="Success", total="total", success=success)

    

    


    


