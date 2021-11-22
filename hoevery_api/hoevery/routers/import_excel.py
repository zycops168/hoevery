from config.db import *
import openpyxl
import pyexcel as p


def import_company(max_col, max_row, sheet_obj):
    success = 0
    _dict = {}

    with SessionContext() as se: # insert into db with _dict
        # insert data to dict (_dict) 
        for i in range(2, max_row + 1): # loop row
            for j in range (1, max_col + 1): # loop column
                cell_obj = sheet_obj.cell(row = i, column = j)
                if j == 1:
                    _dict['name'] = cell_obj.value
                elif j == 2:
                    _dict['description'] = cell_obj.value

            try:
                newcompany = se.query(db.company).filter(db.company.name == _dict['name']).first()
                if newcompany:
                    error = _dict['name'] + " already exits"
                    return dict(ret=-1, msg="Failed", total=max_row-1, success=success, error=dict(row=i-1, detail=error))
            except Exception as e:
                return dict(ret=-1, msg="Not match")

            newcompany = db.company()
            for key in _dict:
                print(newcompany, key)
                if hasattr(newcompany, key):
                    setattr(newcompany, key, _dict.get(key))

            se.add(newcompany)
            se.commit()
            se.refresh(newcompany)
            success += 1
        
    return dict(ret=0, msg="Complete", total=max_row-1, success=success)
    
    

def import_department(max_col, max_row, sheet_obj):
    success = 0
    _dict = {}
    
    with SessionContext() as se: # insert into db with _dict
        for i in range(2, max_row + 1):
            for j in range(1, max_col + 1):
                cel_obj = sheet_obj.cell(row = i , column = j )
                if j == 1:
                    # _dict.setdefault(i-1, {})['corp_id'] = cel_obj.value
                    _dict['company'] = cel_obj.value
                elif j == 2:
                    _dict['name'] = cel_obj.value
                elif j == 3:
                    _dict['description'] = cel_obj.value
            try: 
                newdept = se.query(db.department).filter(db.department.name == _dict['name']).first()
                if newdept:
                    error = 'department name ' + _dict['name'] + ' already exits'
                    return dict(ret=-1, msg="Failed", total=max_row-1, success=success, error=dict(row=i-1, detail=error))
            except Exception as e:
                return dict(ret=-1, msg="Not match")

            try:
                corp_id = se.query(db.company).filter(db.company.name == _dict['company']).first()
                _dict['corp_id']  = int(getattr(corp_id, 'id'))
            except Exception as e:
                return dict(ret=-1, msg=f"company not found at name {_dict['company']}")
                

            newdept = db.department()
            for key in _dict:
                if key == 'company':
                    pass
                elif hasattr(newdept, key):
                    setattr(newdept, key, _dict.get(key))

            se.add(newdept)
            se.commit()
            se.refresh(newdept)
            success += 1

    return dict(ret=0, msg="Complete", total=max_row-1, success=success)
    

def import_employee(max_col, max_row, sheet_obj):
    success = 0
    _dict = {}
    # fake_value
    _dict['create_by'] = "admin1"
    _dict['update_by'] = "admin1"
    _dict['birthdate'] = "2010-01-01" 
    
    with SessionContext() as se: # insert into db with _dict
        for i in range(2, max_row + 1):
            for j in range(1, max_col + 1):
                cel_obj = sheet_obj.cell(row = i , column = j )

                if j == 1:
                    # _dict.setdefault(i-1, {})['corp_id'] = cel_obj.value
                    _dict['company'] = cel_obj.value
                elif j == 2:
                    _dict['department'] = cel_obj.value
                elif j == 3:
                    _dict['firstname'] = cel_obj.value
                elif j == 4:
                    _dict['lastname'] = cel_obj.value
                elif j == 5:
                    _dict['emp_code'] = cel_obj.value
                elif j == 6:
                    _dict['birthDate'] = cel_obj.value
            
            try:
                newemp = se.query(db.employee).filter(db.employee.emp_code == _dict['emp_code']).first()
                if newemp:
                    error = 'emp_code ' + _dict['emp_code'] + ' already exits'
                    return dict(ret=-1, msg="Failed", total=max_row-1, success=success, error=dict(row=i-1, detail=error))
            except Exception as e:
                return dict(ret=-1, msg="Not match")

            try:
                corp_id = se.query(db.company).filter(db.company.name == _dict['company']).first()
                _dict['corp_id']  = int(getattr(corp_id, 'id'))
            except Exception as e:
                return dict(ret=-1, msg=f"company not found at name {_dict['company']}")
            try:
                dept_id = se.query(db.department).filter(db.department.name == _dict['department']).first()
                _dict['dept_id']  = int(getattr(dept_id, 'id'))
            except Exception as e:
                return dict(ret=-1, msg=f"department not found at name {_dict['department']}")
                
            newemp = db.employee()
            for key in _dict:
                if key == 'company' or key == 'department':
                    pass
                elif hasattr(newemp, key):
                    setattr(newemp, key, _dict.get(key))

            se.add(newemp)
            se.commit()
            se.refresh(newemp)
            success += 1

    return dict(ret=0, msg="Complete", total=max_row-1, success=success)


def export_with_pyexcel(file_location):
    _dict = {}
    success = 0
    with SessionContext() as se: # insert into db 
        newcompany = db.company()
        for row in p.get_array(file_name=file_location, start_row=1):
            _dict['company'] = row[0]
            _dict['name'] = row[1]
            _dict['description'] = row[2]
            try:
                newcompany = se.query(db.company).filter(db.company.name == row[1]).first()
                if newcompany:
                    error = row[1] + " already exits"
                    return dict(ret=-1, msg="Failed", total="total", success=success, error=dict(row=success+1, detail=error))
            except Exception as e:
                return dict(ret=-1, msg="Not match")

            newcompany = db.company()
            for key in _dict:
                print(newcompany, key)
                if hasattr(newcompany, key):
                    setattr(newcompany, key, _dict.get(key))

            se.add(newcompany)
            se.commit()
            se.refresh(newcompany)
            success += 1

    return dict(ret=0, msg="Success", total="total", success=success)